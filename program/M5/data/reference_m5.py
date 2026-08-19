"""Эталонные числа модуля M5 (Python).

Запуск (после `generate_m5.py`):

    python reference_m5.py

Пишет рядом с собой четыре эталонных CSV и печатает все числа, на которые
ссылаются шаги модуля:

    ref_load_report.csv     — сколько строк даёт каждый источник (B1)
    ref_join_report.csv     — вход/сопоставлено/потеряно на каждом join (B2)
    ref_partner_report.csv  — итоговый отчёт по партнёрам (B3, B4)
    ref_month_gross.csv     — оборот и комиссия по месяцам (B5)

**Почему эталон написан на стандартной библиотеке, а не на pandas.**
Правило двойного авторства формулы (раздел 1.3 скилла curriculum-design):
то, что учащийся пишет в pandas, и то, чем считается эталон, обязаны быть
разными реализациями. Иначе сверка проверяет, что две копии одного кода
дают одно число, а не что расчёт верен. `openpyxl` здесь используется
только чтобы прочитать ячейки xlsx — разбор шапки и все правила ниже
реализованы отдельно.

**Правила расчёта — канон модуля.** Учащийся сходится с эталоном ровно
тогда, когда воспроизвёл их все; каждое правило названо в шаге, где оно
нужно впервые:

1. Две выгрузки выплат склеиваются. Во второй переименованы две колонки
   (`paid_at` → `payout_date`, `total` → `amount`) и статусы в нижнем
   регистре — статус сравнивается без учёта регистра.
2. В отчёт идут только выплаты со статусом `PAID`.
3. Партнёр сопоставляется по коду; если кода в строке выплаты нет — по
   нормализованному названию (обрезка краёв, схлопывание внутренних
   пробелов, снятие кавычек, приведение регистра через `casefold`).
4. Справочник партнёров дедуплицируется по коду: остаётся первая запись.
5. Валютная выплата пересчитывается по курсу своей даты; если на эту дату
   курса нет (выходной) — по последнему более раннему курсу ряда. Если
   более раннего курса нет вовсе, строка теряется.
6. Выплата без записи о комиссии из отчёта исключается целиком: иначе
   take rate занижается на величину, которую никто не видит.
7. `take_rate` = комиссия / оборот, округление ROUND_HALF_UP до 4 знаков;
   деньги — до 2 знаков.
"""

from __future__ import annotations

import csv
import json
import re
import sys
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

if sys.stdout.encoding is None or sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

HERE = Path(__file__).resolve().parent
RAW = HERE / "raw"

TWO = Decimal("0.01")
FOUR = Decimal("0.0001")
SPACES = re.compile(r"\s+")


def money(value: Decimal) -> Decimal:
    return value.quantize(TWO, rounding=ROUND_HALF_UP)


def norm_name(name: str) -> str:
    """Правило 3. Нормализация названия партнёра."""
    cleaned = name.replace("«", "").replace("»", "").replace('"', "").replace("'", "")
    return SPACES.sub(" ", cleaned).strip().casefold()


def read_payouts() -> list[dict]:
    """Правило 1. Две выгрузки в один список словарей."""
    rows: list[dict] = []
    with (RAW / "payouts_2026_q1.csv").open(encoding="cp1251", newline="") as f:
        for row in csv.DictReader(f, delimiter=";"):
            rows.append({
                "payout_id": int(row["payout_id"]),
                "partner_code": row["partner_code"].strip(),
                "partner_name": row["partner_name"],
                "payout_date": datetime.strptime(row["payout_date"], "%d.%m.%Y").date(),
                "amount": Decimal(row["amount"].replace(",", ".")),
                "currency": row["currency"].strip().upper(),
                "status": row["status"].strip().upper(),
                "source": "payouts_2026_q1.csv",
            })
    with (RAW / "payouts_2026_q2.csv").open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f, delimiter=","):
            rows.append({
                "payout_id": int(row["payout_id"]),
                "partner_code": row["partner_code"].strip(),
                "partner_name": row["partner_name"],
                "payout_date": date.fromisoformat(row["paid_at"]),
                "amount": Decimal(row["total"]),
                "currency": row["currency"].strip().upper(),
                "status": row["status"].strip().upper(),
                "source": "payouts_2026_q2.csv",
            })
    return rows


def read_partners_and_rates() -> tuple[list[dict], list[dict], dict[date, Decimal]]:
    """Лист «Партнери»: шапка занимает две строки, данные начинаются с
    третьей. Правило 4 — дедупликация по коду. Лист «Курс НБУ»: обычная
    шапка в одну строку."""
    wb = load_workbook(RAW / "partners.xlsx", read_only=True, data_only=True)
    raw_rows = []
    for i, row in enumerate(wb["Партнери"].iter_rows(values_only=True), start=1):
        if i <= 2:
            continue  # строка 1 — заголовок выгрузки, строка 2 — имена колонок
        code, name, city, plan, status = (row + (None,) * 5)[:5]
        raw_rows.append({
            "code": "" if code is None else str(code).strip(),
            "name": "" if name is None else str(name),
            "city": "" if city is None else str(city),
            "plan": "" if plan is None else str(plan),
            "status": "" if status is None else str(status),
        })

    partners: list[dict] = []
    seen_codes: set[str] = set()
    for p in raw_rows:
        if p["code"]:
            if p["code"] in seen_codes:
                continue
            seen_codes.add(p["code"])
        partners.append(p)

    rates: dict[date, Decimal] = {}
    for i, row in enumerate(wb["Курс НБУ"].iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        day, value = row[0], row[1]
        rates[date.fromisoformat(str(day))] = Decimal(str(value))
    wb.close()
    return partners, raw_rows, rates


def read_fees() -> tuple[dict[int, Decimal], list[tuple[int, Decimal]], int, int]:
    """Вложенный JSON со страницами. Правило пагинации: дубликаты по
    `payout_id` снимаются, остаётся первая запись. Возвращается и список
    элементов как есть — на нём считается наивный вариант."""
    payload = json.loads((RAW / "fees_api.json").read_text(encoding="utf-8"))
    fees: dict[int, Decimal] = {}
    items: list[tuple[int, Decimal]] = []
    duplicates = 0
    for page in payload["pages"]:
        for item in page["items"]:
            pid = int(item["payout_id"])
            amount = Decimal(item["fee"]["amount"])
            items.append((pid, amount))
            if pid in fees:
                duplicates += 1
                continue
            fees[pid] = amount
    return fees, items, len(items), duplicates


def partner_book(raw_partners: list[dict], drop_dup_codes: bool,
                 collapse_empty: bool) -> list[dict]:
    """Справочник в одном из четырёх состояний. Две развилки независимы:
    что сделано с тремя дублями кода и что сделано с пятью записями без
    кода. Слово «дедуплицирован» само по себе их не различает — и разница
    между «схлопнуть только непустые» и «схлопнуть всё» больше, чем весь
    эффект от дублей."""
    book: list[dict] = []
    seen: set[str] = set()
    for p in raw_partners:
        key = p["code"]
        if key:
            if drop_dup_codes and key in seen:
                continue
        elif collapse_empty and "" in seen:
            continue
        seen.add(key)
        book.append(p)
    return book


def naive_variants(paid: list[dict], raw_partners: list[dict],
                   rates: dict[date, Decimal],
                   fee_items: list[tuple[int, Decimal]]) -> tuple[list[list], dict]:
    """«Наивный merge» — восемь разных чисел, потому что определений восемь.

    Заведено после того, как одно и то же действие дало 5571 у читателя и
    5399 в первой редакции этого файла; третья развилка нашлась таким же
    способом — читатель схлопнул и пустые коды тоже и получил 3523. Все
    числа верны, различаются определения. Развилок три и они независимы:

    1. тип соединения: `inner` (умолчание pandas) или `left`;
    2. три дубля кода в справочнике: оставлены или сняты;
    3. пять записей без кода: оставлены все или схлопнуты в одну.

    Число, попадающее в шаг как «типичная ошибка», обязано приходить со
    всеми тремя ответами рядом — иначе учащийся получит своё и решит, что
    сломал датасет.

    Считается здесь, а не в pandas, по тем же правилам соединения:
    inner — по строке на каждую пару совпавших ключей; left — то же, плюс
    несовпавшая строка остаётся одна. Пустой код совпадает с пустым кодом
    (проверено на pandas 3.0.5: NaN там тоже совпал с NaN)."""
    def rows_after(partners: list[dict], how: str) -> int:
        index: dict[str, int] = {}
        for p in partners:
            index[p["code"]] = index.get(p["code"], 0) + 1
        total = 0
        for p in paid:
            hits = index.get(p["partner_code"], 0)
            total += hits if how == "inner" else max(hits, 1)
        return total

    table = []
    for how in ("inner", "left"):
        for drop_dup, dup_label in ((False, "оставлены"), (True, "сняты")):
            for collapse, empty_label in ((False, "оставлены все пять"),
                                          (True, "схлопнуты в одну")):
                book = partner_book(raw_partners, drop_dup, collapse)
                n = rows_after(book, how)
                table.append([
                    "inner (умолчание pandas)" if how == "inner" else "left",
                    dup_label, empty_label, len(book), len(paid), n,
                    f"{(n / len(paid) - 1) * 100:+.1f}%",
                ])

    # Сквозной наивный путь: справочник как прочитан, merge по умолчанию,
    # дубли пагинации не сняты. Ровно то, что получается, если не сделать
    # ничего из правил 3–6 канона.
    by_code: dict[str, list[dict]] = {}
    for p in raw_partners:
        by_code.setdefault(p["code"], []).append(p)
    fees_by_id: dict[int, list[Decimal]] = {}
    for pid, amount in fee_items:
        fees_by_id.setdefault(pid, []).append(amount)

    rows = [(p, partner) for p in paid for partner in by_code.get(p["partner_code"], [])]
    usd = sum(1 for p, _ in rows if p["currency"] == "USD")
    with_rate = []
    no_rate = 0
    for p, partner in rows:
        if p["currency"] == "UAH":
            with_rate.append((p, Decimal("1")))
            continue
        rate = rate_for(p["payout_date"], rates)
        if rate is None:
            no_rate += 1
            continue
        with_rate.append((p, rate))
    after_fee_join = 0
    gross = fee_total = Decimal("0")
    counted = no_fee = 0
    for p, rate in with_rate:
        matches = fees_by_id.get(p["payout_id"], [])
        after_fee_join += max(len(matches), 1)
        if not matches:
            no_fee += 1
            continue
        for fee in matches:
            counted += 1
            gross += p["amount"] if p["currency"] == "UAH" else money(p["amount"] * rate)
            fee_total += fee
    end_to_end = {
        "после join со справочником": len(rows),
        "валютных строк": usd,
        "без курса": no_rate,
        "после join с комиссиями": after_fee_join,
        "без комиссии": no_fee,
        "строк в итоге": counted,
        "оборот": gross,
        "комиссия": fee_total,
    }
    return table, end_to_end


def rate_for(day: date, rates: dict[date, Decimal]) -> Decimal | None:
    """Правило 5: курс своей даты, иначе последний более ранний."""
    if day in rates:
        return rates[day]
    earlier = [d for d in rates if d < day]
    return rates[max(earlier)] if earlier else None


def main() -> int:
    payouts = read_payouts()
    partners, raw_partners, rates = read_partners_and_rates()
    dup_partners = len(raw_partners) - len(partners)
    fees, fee_items, fees_total, fees_duplicates = read_fees()

    by_code = {p["code"]: p for p in partners if p["code"]}
    by_name: dict[str, dict] = {}
    for p in partners:
        by_name.setdefault(norm_name(p["name"]), p)

    paid = [p for p in payouts if p["status"] == "PAID"]

    # join 1 — партнёр
    matched_partner: list[tuple[dict, dict]] = []
    lost_partner = 0
    for p in paid:
        partner = by_code.get(p["partner_code"]) if p["partner_code"] else \
            by_name.get(norm_name(p["partner_name"]))
        if partner is None:
            lost_partner += 1
            continue
        matched_partner.append((p, partner))

    # join 2 — курс (только валютные строки)
    matched_rate: list[tuple[dict, dict, Decimal]] = []
    lost_rate = 0
    usd_rows = 0
    for p, partner in matched_partner:
        if p["currency"] == "UAH":
            matched_rate.append((p, partner, Decimal("1")))
            continue
        usd_rows += 1
        rate = rate_for(p["payout_date"], rates)
        if rate is None:
            lost_rate += 1
            continue
        matched_rate.append((p, partner, rate))

    # join 3 — комиссия
    final: list[tuple[dict, dict, Decimal, Decimal]] = []
    lost_fee = 0
    for p, partner, rate in matched_rate:
        fee = fees.get(p["payout_id"])
        if fee is None:
            lost_fee += 1
            continue
        final.append((p, partner, rate, fee))

    # --- агрегаты --------------------------------------------------------
    agg: dict[str, dict] = {}
    months: dict[str, dict] = {}
    for p, partner, rate, fee in final:
        gross = p["amount"] if p["currency"] == "UAH" else money(p["amount"] * rate)
        key = partner["code"] or f"name:{norm_name(partner['name'])}"
        row = agg.setdefault(key, {
            "partner_code": partner["code"], "partner_name": partner["name"],
            "city": partner["city"], "plan": partner["plan"],
            "payouts": 0, "gross": Decimal("0"), "fee": Decimal("0"),
        })
        row["payouts"] += 1
        row["gross"] += gross
        row["fee"] += fee
        m = months.setdefault(p["payout_date"].strftime("%Y-%m"),
                              {"gross": Decimal("0"), "fee": Decimal("0")})
        m["gross"] += gross
        m["fee"] += fee

    report = sorted(agg.values(), key=lambda r: (-r["gross"], r["partner_code"]))
    for row in report:
        row["gross"] = money(row["gross"])
        row["fee"] = money(row["fee"])
        row["take_rate"] = (row["fee"] / row["gross"]).quantize(
            FOUR, rounding=ROUND_HALF_UP) if row["gross"] else Decimal("0.0000")

    def write_csv(name: str, header: list[str], rows: list[list]) -> None:
        with (HERE / name).open("w", encoding="utf-8", newline="") as f:
            w = csv.writer(f, lineterminator="\r\n")
            w.writerow(header)
            w.writerows(rows)

    q1 = sum(1 for p in payouts if p["source"] == "payouts_2026_q1.csv")
    q2 = len(payouts) - q1
    write_csv("ref_load_report.csv", ["источник", "строк"], [
        ["payouts_2026_q1.csv", q1],
        ["payouts_2026_q2.csv", q2],
        ["выплаты всего", len(payouts)],
        ["выплаты со статусом PAID", len(paid)],
        ["partners.xlsx, лист «Партнери» (строк данных)", len(partners) + dup_partners],
        ["partners.xlsx, после дедупликации по коду", len(partners)],
        ["partners.xlsx, лист «Курс НБУ»", len(rates)],
        ["fees_api.json, элементов на страницах", fees_total],
        ["fees_api.json, после снятия дублей пагинации", len(fees)],
    ])

    write_csv("ref_join_report.csv",
              ["join", "вход", "сопоставлено", "потеряно", "причина потери"], [
        ["выплаты PAID × справочник партнёров", len(paid), len(matched_partner),
         lost_partner, "кода нет в справочнике (партнёра подключили, выгрузку CRM не обновили)"],
        ["выплаты × курс НБУ", len(matched_partner), len(matched_rate), lost_rate,
         "валютная выплата раньше начала ряда курсов: нет курса ни на дату, ни ранее"],
        ["выплаты × комиссии", len(matched_rate), len(final), lost_fee,
         "комиссия по выплате не рассчитана — записи в API нет"],
    ])

    write_csv("ref_partner_report.csv",
              ["partner_code", "partner_name", "city", "plan", "payouts",
               "gross_uah", "fee_uah", "take_rate"],
              [[r["partner_code"], r["partner_name"], r["city"], r["plan"],
                r["payouts"], f"{r['gross']:.2f}", f"{r['fee']:.2f}",
                f"{r['take_rate']:.4f}"] for r in report])

    write_csv("ref_month_gross.csv", ["month", "gross_uah", "fee_uah"],
              [[m, f"{months[m]['gross']:.2f}", f"{months[m]['fee']:.2f}"]
               for m in sorted(months)])

    naive_table, naive_end = naive_variants(paid, raw_partners, rates, fee_items)
    # Совпадение чисел опаснее расхождения: вариант «дубли сняты, записи без
    # кода схлопнуты в одну» даёт ровно столько же строк, сколько эталон.
    # Строки при этом другие — выплаты без кода получают одного и того же
    # партнёра вместо своего по названию.
    collapsed_book = partner_book(raw_partners, True, True)
    stub = next(b for b in collapsed_book if not b["code"])
    wrong_partner = sum(1 for pay, partner in matched_partner
                        if not pay["partner_code"] and partner is not stub)
    # Сколько выплат заглушка «угадала» — зависит от того, как сравниваются
    # названия. Три сравнения — три числа; в тексте стоит третье, но и два
    # других воспроизводимы, чтобы читатель со своим числом нашёл себя.
    codeless = [pay for pay in paid if not pay["partner_code"]]
    stub_hits = {
        "строгое равенство": sum(1 for pay in codeless
                                 if pay["partner_name"] == stub["name"]),
        "одна обрезка краёв": sum(1 for pay in codeless
                                  if pay["partner_name"].strip() == stub["name"].strip()),
        "правило 3 канона": sum(1 for pay in codeless
                                if norm_name(pay["partner_name"]) == norm_name(stub["name"])),
    }
    write_csv("ref_naive_merge.csv",
              ["способ соединения", "дубли кода", "записи без кода",
               "строк в справочнике", "вход", "строк после merge", "прирост"],
              naive_table)

    # --- печать всех числ, на которые будут ссылаться шаги ---------------
    print("B1 — загрузка четырёх файлов")
    print(f"  payouts_2026_q1.csv: {q1} строк")
    print(f"  payouts_2026_q2.csv: {q2} строк")
    print(f"  всего выплат после склейки: {len(payouts)}")
    print(f"  из них PAID: {len(paid)}")
    print(f"  справочник партнёров: {len(partners) + dup_partners} строк данных, "
          f"дублей по коду {dup_partners}, после дедупликации {len(partners)}")
    print(f"  без кода в справочнике: {sum(1 for p in partners if not p['code'])}")
    print(f"  курс НБУ: {len(rates)} рабочих дней")
    print(f"  комиссии: {fees_total} элементов на {len(json.loads((RAW / 'fees_api.json').read_text(encoding='utf-8'))['pages'])} страницах, "
          f"дублей пагинации {fees_duplicates}, уникальных {len(fees)}")
    print()
    print("B2 — потери на join'ах")
    print(f"  выплаты PAID × справочник: вход {len(paid)}, сопоставлено "
          f"{len(matched_partner)}, потеряно {lost_partner}")
    print(f"  выплаты × курс: вход {len(matched_partner)}, сопоставлено "
          f"{len(matched_rate)}, потеряно {lost_rate} (валютных строк {usd_rows})")
    print(f"  выплаты × комиссии: вход {len(matched_rate)}, сопоставлено "
          f"{len(final)}, потеряно {lost_fee}")
    print(f"  арифметика сходится: "
          f"{len(paid) == len(final) + lost_partner + lost_rate + lost_fee}")
    print()
    print("B3, B4 — итоговый отчёт")
    total_gross = sum(r["gross"] for r in report)
    total_fee = sum(r["fee"] for r in report)
    print(f"  партнёров в отчёте: {len(report)}")
    print(f"  оборот всего: {total_gross:.2f} грн")
    print(f"  комиссия всего: {total_fee:.2f} грн")
    print(f"  take rate по всему портфелю: "
          f"{(total_fee / total_gross).quantize(FOUR, rounding=ROUND_HALF_UP):.4f}")
    top = report[0]
    print(f"  первая строка отчёта: {top['partner_name']} "
          f"({top['partner_code'] or 'без кода'}), {top['payouts']} выплат, "
          f"{top['gross']:.2f} грн, take rate {top['take_rate']:.4f}")
    print()
    print("B5 — по месяцам")
    for m in sorted(months):
        print(f"  {m}: оборот {months[m]['gross']:.2f}, комиссия {months[m]['fee']:.2f}")

    print()
    print(f"«Наивный merge» — восемь определений, восемь чисел (вход {len(paid)} PAID)")
    for how, dup, empty, book_n, _, rows_n, growth in naive_table:
        print(f"  {how:<24} дубли {dup:<10} без кода {empty:<18} "
              f"справочник {book_n} -> {rows_n} строк, {growth}")
    print(f"  вариант «дубли сняты, без кода схлопнуты» даёт столько же строк, "
          f"сколько эталон ({len(matched_partner)}), но у {wrong_partner} из них "
          f"партнёр не тот")
    print(f"  заглушка при схлопывании — {stub['name']} (первая по порядку строк);"
          f" из {len(codeless)} выплат без кода она угадала:")
    for how, n in stub_hits.items():
        print(f"    {how}: {n}")
    print("  сквозной наивный путь (справочник как прочитан, merge по умолчанию,")
    print("  дубли пагинации не сняты):")
    for key, value in naive_end.items():
        printed = f"{value:.2f}" if isinstance(value, Decimal) else value
        print(f"    {key}: {printed}")
    print(f"    завышение оборота против эталона: "
          f"{(naive_end['оборот'] / total_gross - 1) * 100:.1f}%")
    print(f"    завышение комиссии против эталона: "
          f"{(naive_end['комиссия'] / total_fee - 1) * 100:.1f}%")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
